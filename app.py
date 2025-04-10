import os, re, time, threading
from flask import Flask, request, render_template, send_file, jsonify
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

progress = {
    "total": 0,
    "current": 0,
    "status": ""
}

@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    if not file:
        return "No file uploaded."
    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    thread = threading.Thread(target=run_bot, args=(path,))
    thread.start()

    return jsonify({"message": "File uploaded and processing started."})

@app.route('/progress')
def get_progress():
    return jsonify(progress)

def run_bot(excel_path):
    wb = load_workbook(excel_path)
    sh = wb.active
    driver = webdriver.Chrome()
    driver.maximize_window()

    progress['total'] = sh.max_row - 1
    progress['current'] = 0
    progress['status'] = 'Processing'

    for i in range(2, sh.max_row + 1):
        try:
            data = [sh[f'{col}{i}'].value for col in 'ABCDE']
            if not all(data):
                sh[f'F{i}'] = '❌ Missing data'
                continue

            url, title, desc, name, email = data

            driver.get("https://ebay-dir.com/submit?c=51&LINK_TYPE=1")
            time.sleep(2)

            driver.find_element(By.ID, 'TITLE').send_keys(title)
            time.sleep(0.5)
            driver.find_element(By.ID, 'URL').send_keys(url)
            time.sleep(0.5)
            driver.find_element(By.ID, 'DESCRIPTION').send_keys(desc)
            time.sleep(0.5)
            driver.find_element(By.ID, 'OWNER_NAME').send_keys(name)
            time.sleep(0.5)
            driver.find_element(By.ID, 'OWNER_EMAIL').send_keys(email)
            time.sleep(0.5)

            for f in driver.find_elements(By.TAG_NAME, 'font'):
                if '=' in f.text:
                    expr = f.text.replace('x', '*').replace('=', '')
                    driver.find_element(By.ID, "DO_MATH").send_keys(str(eval(expr)))
                    break

            driver.find_element(By.XPATH, "//input[@id='AGREERULES']").click()
            driver.find_element(By.NAME, 'continue').click()
            time.sleep(2)

            body = driver.find_element(By.TAG_NAME, 'body').text
            link = re.search(r'https?://\S+\.html', body)
            sh[f'F{i}'] = '✅ Posted'
            sh[f'G{i}'] = link.group() if link else 'No link'
        except Exception as e:
            sh[f'F{i}'] = '❌ Failed'
        progress['current'] += 1

    driver.quit()
    wb.save(excel_path)
    progress['status'] = 'Completed'

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(UPLOAD_FOLDER, filename), as_attachment=True)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5050))
    app.run(host='0.0.0.0', port=port)
